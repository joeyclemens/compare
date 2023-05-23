import tkinter as tk
from tkinter import ttk, messagebox
import csv
from tkinter import filedialog
from tkinter.filedialog import askopenfilename, asksaveasfilename
import os
import openpyxl

# Create the main window
root = tk.Tk()
root.geometry("500x400")
root.title("File Comparator")
# Create a style object
style = ttk.Style()
# Set the background color of the window
root.configure(bg="#d4d4d4")
# Set the theme to 'clam'
style.theme_use('classic')



# Create the function to select the input Excel files
def select_input_file(label):
    file_paths = filedialog.askopenfilenames(filetypes=[('Excel Files', '*.xlsx')])
    if not file_paths:
        return ''
    label.config(text=', '.join(os.path.basename(file_path) for file_path in file_paths))
    return ';'.join(os.path.abspath(file_path) for file_path in file_paths)

# Create the function to select the output Excel file
def select_output_file():
    file_path = asksaveasfilename(defaultextension=".xlxs", filetypes=[('Excel Files', '*.xlsx')])
    if file_path == '':
        return ''
    return file_path

# Create the label and entry for the first Excel file
file1_frame = tk.Frame(root, bg="#d4d4d4")
file1_frame.pack(pady=10)

file1_label = tk.Label(file1_frame, text="Select the first file:", bg="#d4d4d4")
file1_label.pack(side=tk.LEFT, padx=(0, 10))

file1_entry = ttk.Entry(file1_frame)
file1_entry.pack(side=tk.LEFT, fill=tk.X, expand=True)

file1_button = tk.Button(file1_frame, text="Browse", command=lambda: file1_entry.insert(tk.END, select_input_file(file1_name_label)))
file1_button.pack(side=tk.LEFT)
# Create the label and entry for the second file
file2_frame = tk.Frame(root, bg="#d4d4d4")
file2_frame.pack(pady=10)

file2_label = tk.Label(file2_frame, text="Select the second file:", bg="#d4d4d4")
file2_label.pack(side=tk.LEFT, padx=(0, 10))

file2_entry = ttk.Entry(file2_frame)
file2_entry.pack(side=tk.LEFT, fill=tk.X, expand=True)

file2_button = tk.Button(file2_frame, text="Browse", command=lambda: file2_entry.insert(tk.END, select_input_file(file2_name_label)))
file2_button.pack(side=tk.LEFT)

# Create the label and drop-down menu for the matching columns
match_frame = tk.Frame(root, bg="#d4d4d4")
match_frame.pack(pady=10)

#match_label = tk.Label(match_frame, text="Select the column or columns to match on:", bg="#d4d4d4")
#match_label.pack(side=tk.LEFT, padx=(0, 10))

match_listbox = tk.Listbox(match_frame, selectmode=tk.MULTIPLE)
match_listbox.pack(side=tk.LEFT)

# Create the function to update the matching column options
def update_match_options(event=None):
    file1_path = file1_entry.get()
    file2_path = file2_entry.get()

    if file1_path and file2_path:
        headers1 = get_headers(file1_path)
        headers2 = get_headers(file2_path)
        all_headers = list(set(headers1 + headers2)) # combine headers from both files and remove duplicates
        all_headers.sort() # sort the headers alphabetically
        match_listbox.delete(0, tk.END)
        for col in all_headers:
            match_listbox.insert(tk.END, col)
                
# Create the Refresh button
refresh_button = tk.Button(match_frame, text="Select the column or columns to match on", command=update_match_options)
refresh_button.pack(side=tk.LEFT, padx=(10, 0))
                

file1_entry.bind("<FocusOut>", update_match_options)
file2_entry.bind("<FocusOut>", update_match_options)


# Create the label and entry for the output file
output_frame = tk.Frame(root, bg="#d4d4d4")
output_frame.pack(pady=10)

output_label = tk.Label(output_frame, text="Select the output file:", bg="#d4d4d4")
output_label.pack(side=tk.LEFT, padx=(0, 10))

output_entry = ttk.Entry(output_frame)
output_entry.pack(side=tk.LEFT, fill=tk.X, expand=True)

output_button = tk.Button(output_frame, text="Browse", command=lambda: output_entry.insert(tk.END, select_output_file()))
output_button.pack(side=tk.LEFT)


# Create the function to get the headers of the files
def get_headers(file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    headers = []
    for cell in sheet[1]:
        headers.append(cell.value)
    return headers



# Create the function to compare the files
def compare_csv():
    file1_path = file1_entry.get()
    file2_path = file2_entry.get()
    output_path = output_entry.get()
    matching_cols = match_listbox.curselection()
    if not file1_path or not file2_path or not output_path or not matching_cols:
        messagebox.showerror("Error", "Please select both Excel files, the output file, and the matching columns.")
        return
    matching_cols = [match_listbox.get(idx) for idx in matching_cols]

    headers1 = get_headers(file1_path)
    headers2 = get_headers(file2_path)
    all_headers = headers1 + headers2

    if any(col not in all_headers for col in matching_cols):
        messagebox.showerror("Error", "One or more matching columns not found in the input Excel files.")
        return

    try:
        wb1 = openpyxl.load_workbook(file1_path)
        sheet1 = wb1.active
        wb2 = openpyxl.load_workbook(file2_path)
        sheet2 = wb2.active

        df1 = []
        for row in sheet1.iter_rows(min_row=2, values_only=True):
            df1.append(dict(zip(headers1, row)))
        df2 = []
        for row in sheet2.iter_rows(min_row=2, values_only=True):
            df2.append(dict(zip(headers2, row)))

        matching_rows = []
        non_matching_rows1 = []
        non_matching_rows2 = []
        for row1 in df1:
            match_found = False
            for row2 in df2:
                if all(row1[col] == row2[col] for col in matching_cols):
                    matching_row = {**row1, **row2}
                    matching_rows.append(matching_row)
                    match_found = True
                    break
            if not match_found:
                non_matching_rows1.append(row1)
        for row2 in df2:
            match_found = False
            for row1 in df1:
                if all(row1[col] == row2[col] for col in matching_cols):
                    match_found = True
                    break
            if not match_found:
                non_matching_rows2.append(row2)

        wb3 = openpyxl.Workbook()
        sheet3 = wb3.active
        sheet3.append(all_headers)

        for matching_row in matching_rows:
            output_row = []
            for header in all_headers:
                output_row.append(matching_row.get(header, ''))
            sheet3.append(output_row)

        for non_matching_row in non_matching_rows1:
            output_row = []
            for header in all_headers:
                output_row.append(non_matching_row.get(header, ''))
            output_row[len(headers1):] = [''] * len(headers2)
            sheet3.append(output_row)

        for non_matching_row in non_matching_rows2:
            output_row = [''] * len(headers1)
            for header in headers2:
                output_row.append(non_matching_row.get(header, ''))
            sheet3.append(output_row)

        wb3.save(output_path)

        messagebox.showinfo("Success", "The new Excel file has been created.")
        # Open the output file
        os.startfile(output_path)
    except:
        messagebox.showerror("Error", "An error occurred while processing the Excel files. Are the column headers a 1:1 match?")
    
def show_text():
    # Create a new window
    popup = tk.Toplevel(root)
    popup.title("To be honest, I dont know what im doing either")
    
    # Create a label with the text to display
    text_label = tk.Label(popup, text="This program is designed to read and write data from Excel files. \nTo use it, you will need to input the data you want to compare into two separate Excel files.\n Please note that there is a minor bug/issue in the program where the columns you select to make the match with\n need to have the same column header name.\n The program should run smoothly as long as you keep this in mind.\n If you are comparing a (SoL), it is recommended to scroll to the bottom of your output file to where the non matching rows on the right side are\n and copy the roomId headers for these to the corresponding column on the left. \nThen, sort the column alphabetically to ensure that the data is properly matched. \nJust ask Joey if you need help with this")
    text_label.pack()
    

    # Create a button to close the window
    close_button = tk.Button(popup, text="Close", command=popup.destroy)
    close_button.pack()
    
# Create a separator widget
separator = ttk.Separator(root, orient="horizontal")
separator.pack(fill="x")    
    
# Create a button to show the text
text_button = ttk.Button(root, text="Help", command=show_text)
text_button.pack(side="left")    
    
# Create the frame for displaying the selected file names
file_frame = tk.Frame(root, bg="#d4d4d4")
file_frame.pack(side=tk.BOTTOM, pady=10)    

# Create the label for the first file name
file1_name_label = tk.Label(file_frame, text="", bg="#d4d4d4")
file1_name_label.pack(side=tk.LEFT, padx=(0, 10))

# Create the label for the second file name
file2_name_label = tk.Label(file_frame, text="", bg="#d4d4d4")
file2_name_label.pack(side=tk.LEFT)

# Create the button to compare the files
compare_button = tk.Button(root, text="Compare Your Files", command=compare_csv)
compare_button.pack(pady=10)



# Run the main loop
root.mainloop()    
